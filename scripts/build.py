#!/usr/bin/env python3
"""Build trip_report.xlsx + trip_dashboard.html (+ public) from data/*.json."""

from __future__ import annotations

import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUT = ROOT
TEMPLATE = ROOT / "templates" / "dashboard.html"
VENV_PY = ROOT / ".venv" / "bin" / "python"

# Card last-4 / account numbers must not ship in the public dashboard.
_RE_CARD = re.compile(r"\s*\*\d{2,}")
_RE_ACCT = re.compile(r"(счёт\s+NBG)\s+\d+", re.IGNORECASE)


def _ensure_venv() -> None:
    """Code Runner / system python3 often lack openpyxl — re-exec via .venv."""
    if os.environ.get("TRIP_BUILD_VENV") == "1":
        return
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass
    if VENV_PY.is_file():
        os.environ["TRIP_BUILD_VENV"] = "1"
        os.execv(str(VENV_PY), [str(VENV_PY), str(Path(__file__).resolve()), *sys.argv[1:]])
    print(
        "Need openpyxl. From project root:\n"
        "  python3 -m venv .venv && .venv/bin/pip install -r requirements.txt\n"
        "Then re-run, or:  .venv/bin/python scripts/build.py",
        file=sys.stderr,
    )
    sys.exit(1)


_ensure_venv()

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def load() -> tuple[list, list, dict]:
    required = {
        DATA / "transactions.json": None,  # committed
        DATA / "excluded.json": "excluded.example.json",
        DATA / "meta.json": "meta.example.json",
    }
    missing = [p for p in required if not p.is_file()]
    if missing:
        lines = ["Missing local data files (gitignored). From project root:"]
        for p in missing:
            example = required[p]
            if example:
                lines.append(f"  cp data/{example} {p.relative_to(ROOT)}")
            else:
                lines.append(f"  (restore) {p.relative_to(ROOT)}")
        lines.append("Then re-run:  .venv/bin/python scripts/build.py")
        print("\n".join(lines), file=sys.stderr)
        sys.exit(1)
    txs = _read_json(DATA / "transactions.json")
    excl = _read_json(DATA / "excluded.json")
    meta = _read_json(DATA / "meta.json")
    return txs, excl, meta


def public_src(src: str) -> str:
    """Strip card last-4 and account numbers from a payment-source label."""
    s = _RE_CARD.sub("", src)
    s = _RE_ACCT.sub(r"\1", s)
    return s.strip() or "NBG card"


def scrub_public(txs: list, meta: dict) -> tuple[list, dict]:
    """Copy of txs/meta safe to embed in the shareable HTML."""
    txs_pub = [{**t, "src": public_src(t["src"])} for t in txs]
    meta_pub = json.loads(json.dumps(meta))
    meta_pub.pop("account_picture", None)
    meta_pub.pop("financing", None)
    meta_pub["sources"] = ["Revolut", "PayPal", "NBG cards"]
    meta_pub["footer"] = (
        "Публичная версия: без деталей банковского счёта и номеров карт. "
        "Отменённые брони и внутренние переводы исключены из сумм."
    )
    cs = meta_pub.get("checksum") or {}
    by_src = cs.get("by_source") or {}
    merged: dict[str, float] = defaultdict(float)
    for k, v in by_src.items():
        merged[public_src(k)] += float(v)
    cs["by_source"] = {k: round(v, 2) for k, v in sorted(merged.items())}
    meta_pub["checksum"] = cs
    return txs_pub, meta_pub


def checksum(txs: list, travelers: int = 3) -> dict:
    n = max(int(travelers or 3), 1)
    total = -sum(t["amount_eur"] for t in txs)
    pre = -sum(t["amount_eur"] for t in txs if t.get("pretrip"))
    shared = -sum(t["amount_eur"] for t in txs if not t.get("personal"))
    personal = -sum(t["amount_eur"] for t in txs if t.get("personal"))
    by_src: dict[str, float] = defaultdict(float)
    for t in txs:
        by_src[t["src"]] -= t["amount_eur"]
    return {
        "trip_total": round(total, 2),
        "pretrip": round(pre, 2),
        "in_trip": round(total - pre, 2),
        "shared": round(shared, 2),
        "personal": round(personal, 2),
        "per_person": round(shared / n, 2),
        "by_source": {k: round(v, 2) for k, v in sorted(by_src.items())},
        "n": len(txs),
    }


def verify(txs: list, meta: dict) -> None:
    n_trav = int(meta.get("travelers") or 3)
    cs = checksum(txs, n_trav)
    expected = meta.get("checksum", {})
    mismatches = []
    for key in ("trip_total", "pretrip", "in_trip", "shared", "personal", "per_person"):
        if key in expected and round(expected[key], 2) != cs[key]:
            mismatches.append(f"{key}: meta={expected[key]} actual={cs[key]}")
    if mismatches:
        print("CHECKSUM MISMATCH — updating meta.json checksum from data:")
        for m in mismatches:
            print(" ", m)
        meta["checksum"] = cs
        (DATA / "meta.json").write_text(
            json.dumps(meta, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
    else:
        print(f"OK checksum: {cs['trip_total']} € · {cs['n']} txs · per person {cs['per_person']} €")
    # always refresh computed checksum in meta for embed
    meta["checksum"] = cs


def build_html(txs: list, excl: list, meta: dict, public: bool) -> Path:
    tpl = TEMPLATE.read_text(encoding="utf-8")
    if public:
        txs_emb, meta_emb = scrub_public(txs, meta)
        excl_emb: list = []
    else:
        txs_emb, meta_emb, excl_emb = txs, meta, excl
    html = (
        tpl.replace("__DATA__", json.dumps(txs_emb, ensure_ascii=False))
        .replace("__EXCLUDED__", json.dumps(excl_emb, ensure_ascii=False))
        .replace("__META__", json.dumps(meta_emb, ensure_ascii=False))
        .replace("__PUBLIC__", "true" if public else "false")
    )
    if public:
        out = OUT / "trip_dashboard_public.html"
        out.write_text(html, encoding="utf-8")
        print(f"Wrote {out.relative_to(ROOT)} ({out.stat().st_size // 1024} KB)")
        # GitHub Pages
        docs = OUT / "docs"
        docs.mkdir(exist_ok=True)
        (docs / ".nojekyll").write_text("", encoding="utf-8")
        index = docs / "index.html"
        index.write_text(html, encoding="utf-8")
        print(f"Wrote {index.relative_to(ROOT)} ({index.stat().st_size // 1024} KB)")
        return out
    out = OUT / "trip_dashboard.html"
    out.write_text(html, encoding="utf-8")
    print(f"Wrote {out.relative_to(ROOT)} ({out.stat().st_size // 1024} KB)")
    return out


def _autosize(ws, widths: dict[int, int] | None = None) -> None:
    for i, col in enumerate(ws.columns, 1):
        if widths and i in widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i]
            continue
        maxlen = 0
        for cell in col:
            if cell.value is not None:
                maxlen = max(maxlen, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(i)].width = max(10, maxlen + 2)


def _header(ws, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="1C2333")
    font = Font(color="FFFFFF", bold=True)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def build_xlsx(txs: list, excl: list, meta: dict) -> Path:
    wb = Workbook()

    # --- Транзакции ---
    ws = wb.active
    ws.title = "Транзакции"
    headers = [
        "Дата",
        "Источник",
        "Описание (выписка)",
        "Подпись",
        "Категория",
        "Этап",
        "Оплачено",
        "Личное",
        "День трипа",
        "Сумма, €",
    ]
    _header(ws, headers)
    for i, t in enumerate(sorted(txs, key=lambda x: x["dt"]), 2):
        ws.cell(i, 1, t["dt"][:16].replace("T", " "))
        ws.cell(i, 2, t["src"])
        ws.cell(i, 3, t["desc"])
        ws.cell(i, 4, t["label"])
        ws.cell(i, 5, t["category"])
        ws.cell(i, 6, t["stage"])
        ws.cell(i, 7, "пре-трип" if t.get("pretrip") else "в поездке")
        ws.cell(i, 8, "да" if t.get("personal") else "")
        ws.cell(i, 9, t.get("trip_day") or "")
        ws.cell(i, 10, t["amount_eur"])
    _autosize(ws, {3: 40, 4: 50, 10: 12})

    # --- Сводка ---
    ws = wb.create_sheet("Сводка")
    stages = meta.get("stages") or sorted({t["stage"] for t in txs})
    _header(ws, ["Категория", *stages, "Итого"])
    cats = sorted({t["category"] for t in txs})
    by = defaultdict(float)
    for t in txs:
        by[(t["category"], t["stage"])] += t["amount_eur"]
        by[(t["category"], "_total")] += t["amount_eur"]
        by[("_total", t["stage"])] += t["amount_eur"]
        by[("_total", "_total")] += t["amount_eur"]
    for r, cat in enumerate(cats, 2):
        ws.cell(r, 1, cat)
        for c, st in enumerate(stages, 2):
            ws.cell(r, c, round(by[(cat, st)], 2) or 0)
        ws.cell(r, len(stages) + 2, round(by[(cat, "_total")], 2))
    r = len(cats) + 2
    ws.cell(r, 1, "ИТОГО").font = Font(bold=True)
    for c, st in enumerate(stages, 2):
        ws.cell(r, c, round(by[("_total", st)], 2)).font = Font(bold=True)
    ws.cell(r, len(stages) + 2, round(by[("_total", "_total")], 2)).font = Font(bold=True)

    cs = meta["checksum"]
    n_trav = int(meta.get("travelers") or 3)
    ws.cell(r + 2, 1, "Оплачено до поездки (пре-трип)")
    ws.cell(r + 2, 2, -cs["pretrip"])
    ws.cell(r + 3, 1, "Оплачено в поездке")
    ws.cell(r + 3, 2, -cs["in_trip"])
    ws.cell(r + 4, 1, f"Shared (делится на {n_trav})")
    ws.cell(r + 4, 2, -cs["shared"])
    ws.cell(r + 5, 1, "Личные (не делятся)")
    ws.cell(r + 5, 2, -cs["personal"])
    ws.cell(r + 6, 1, f"На человека (shared ÷ {n_trav})")
    ws.cell(r + 6, 2, -cs["per_person"])
    ws.cell(r + 8, 1, meta.get("period", ""))
    _autosize(ws)

    # --- Исключено ---
    ws = wb.create_sheet("Исключено")
    _header(ws, ["Дата", "Источник", "Описание (выписка)", "Причина исключения", "Сумма", "Валюта"])
    for i, e in enumerate(excl, 2):
        ws.cell(i, 1, str(e["dt"])[:16].replace("T", " "))
        ws.cell(i, 2, e["src"])
        ws.cell(i, 3, e["desc"])
        ws.cell(i, 4, e["reason"])
        ws.cell(i, 5, e["amount"])
        ws.cell(i, 6, e.get("currency") or "EUR")
    _autosize(ws, {3: 40, 4: 45})

    # --- Полная картина (private) ---
    ws = wb.create_sheet("Полная картина")
    ap = meta.get("account_picture") or {}
    ws.cell(1, 1, ap.get("title", "Полная картина")).font = Font(bold=True, size=12)
    ws.cell(3, 1, "Баланс на 02.05.2026")
    ws.cell(3, 2, ap.get("balance_from"))
    ws.cell(4, 1, "Баланс на 02.08.2026")
    ws.cell(4, 2, ap.get("balance_to"))
    ws.cell(6, 1, "Движения").font = Font(bold=True)
    for i, (label, val) in enumerate(ap.get("rows") or [], 7):
        ws.cell(i, 1, label)
        ws.cell(i, 2, val)
    row = 7 + len(ap.get("rows") or []) + 2
    fin = meta.get("financing") or {}
    ws.cell(row, 1, fin.get("headline", "")).font = Font(bold=True)
    for j, line in enumerate(fin.get("lines") or [], 1):
        ws.cell(row + j, 1, "· " + line)
    ws.cell(row + len(fin.get("lines") or []) + 2, 1, fin.get("cash_note", ""))
    _autosize(ws, {1: 70, 2: 14})

    out = OUT / "trip_report.xlsx"
    # if Excel has the file locked, write alternate then warn
    try:
        wb.save(out)
    except PermissionError:
        alt = OUT / "trip_report_built.xlsx"
        wb.save(alt)
        print(f"WARN: {out.name} locked — wrote {alt.name}")
        return alt
    print(f"Wrote {out.relative_to(ROOT)}")
    return out


def main() -> None:
    txs, excl, meta = load()
    verify(txs, meta)
    build_html(txs, excl, meta, public=False)
    build_html(txs, excl, meta, public=True)
    build_xlsx(txs, excl, meta)
    print("Done", datetime.now().isoformat(timespec="seconds"))


if __name__ == "__main__":
    main()
